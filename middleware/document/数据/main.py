from openpyxl import load_workbook
from html.parser import HTMLParser
import requests,base64,os
from xml.etree import ElementTree  as ET
# 运行之前记得关闭DevSidecar或者别的代理工具


# 从1、GetLisRequest 接口（获取标本信息）
def get_patient_info(hospSampleID):
    data = f'''<?xml version="1.0" encoding="utf-8"?>
            <soap12:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap12="http://www.w3.org/2003/05/soap-envelope">
            <soap12:Body><GetLisRequest xmlns="http://tempuri.org/"><hospSampleID>{hospSampleID}</hospSampleID>
            </GetLisRequest></soap12:Body></soap12:Envelope>'''
    encode_data = data.encode('utf-8')
    # print(data)
    headers = {"Host": "10.10.11.196",
            "Content-Type": "application/soap+xml; charset=UTF-8",
            "Content-Length": str(len(encode_data)),
            "SOAPAction": "http://tempuri.org/GetLisRequest"}
    patient_info = requests.post('http://58.62.17.237:4431/ExtReportService', data=encode_data ,headers=headers)

    print('patient_info.header----')
    print(patient_info.headers)
    print('patient_info.text----') 
    print(patient_info.text)

    return patient_info.text


class Parse(HTMLParser):
    def __init__(self):
    #Since Python 3, we need to call the __init__() function 
    #of the parent class
        super().__init__()
        self.reset()
        self.data = None
 
    def handle_data(self, data):
        # print("Here's the data: ", data)
        print('HTMLParser parsering data ')
        self.data = data

 
 # 解析1GetLisRequest返回的xml信息：
def get_LisRequest_data(patient_info):
    def get_tag_raw(tag_name):
        return root[0].find(tag_name).text    
    dct = {}
    parser = Parse()
    parser.feed(patient_info)
    data = parser.data
    root = ET.fromstring(data)
    
    dct['lis_Barcode'] = get_tag_raw('lis_Barcode')
    dct['blood_time'] = get_tag_raw('blood_time') 
    dct['lis_combine_code'] = root[0].find('LisItems').find('lis_item_code').text
        
    all_sub_items = root[0].find('LisItems').findall('SubItems') 
    dct['all_lis_item_code'] =  [item.find('lis_subitem_code').text for item in all_sub_items]
    dct['all_lis_item_name'] =  [item.find('lis_subitem_name').text for item in all_sub_items]    
       
    print('----patient_info-all_lis_item_name dct -----') 
    print(dct)
    # print(all_lis_item_name)
    # with open('xmlfile/'+dct['lis_Barcode'] +'.xml','w',encoding='utf-8') as fh:
    #     fh.write(data)
    # return lis_Barcode,blood_time,lis_combine_code,all_lis_item_code
    return dct
   

# 解析Excel报告的信息,这里的flag有两个，分别是5和17，
# 如果是5，那就是 “雄激素5项(爱湾)”，如果是17，那就代表“17α-羟孕酮”
def get_excel_info(infile,hospSampleID,flag = '5'):
    all_lis_item_code =  ['32378', '32379', '32380', '32381', '32383']
    patient_info = get_patient_info(hospSampleID)
    dct = get_LisRequest_data(patient_info)
    ex  = load_workbook(infile)
    tb = ex['报告']
    lis_Barcode= tb.cell(row=4,column=3).value
    if flag == '5':
        dct['lis_Barcode']   = lis_Barcode.split('+')[0]
        
    if flag == '17':
        dct['lis_Barcode']   = lis_Barcode.split('+')[1]
    # 采样日期： 
    dct['blood_time'] = tb.cell(row=6,column=3).value
    dct['ext_receive_time'],dct['ext_check_time'],dct['ext_report_time'] =  tb.cell(row=7,column=3).value, \
                                                                            tb.cell(row=7,column=7).value, \
                                                                            tb.cell(row=7,column=10).value
   # sheet里面从B10到H16,即检测结果的内容：
    result_info = ''
    result_dct = {}
    i  = 0
    all_lis_item_code =  ['32378', '32379', '32380', '32383', '32381','32382']
    all_lis_item_name =  ['睾酮', '雄烯二酮', '硫酸脱氢表雄酮', '双氢睾酮', '脱氢表雄酮','17α-羟孕酮']
    all_lis_combine_code = ['32025']*5 + ['32024']
    all_ref = [
        "0.02-0.45",
        ' 0.30-2.35',
        '280-3900',
        '＜0.30',
        '19-30岁<13\n31-40岁<10\n41-50岁<8',
        '卵泡期0.09-2.00;黄体期0.3-2.34'
    ]
    result_unit = 'ng/mL'
    for row in range(10,16):
        i += 1
        if flag == '5' and i ==6:
            continue 
        if flag == '17' and i  in (1,2,3,4,5):
            continue 
#         result_dct['ext_item_code'] = all_lis_item_code
        result_dct['ext_item_code'] = all_lis_item_code[i-1]
        result_dct['lis_item_name'] = all_lis_item_name[i-1]
        # ext_item_name = tb.cell(row=row,column=2).value
        result_dct['lis_item_code'] = all_lis_item_code[i-1] 
        result_dct['result_reference'] = all_ref[i-1] 
        result_dct['result_unit'] = result_unit 
        result_dct['lis_combine_code'] = all_lis_combine_code[i-1]  

        result_dct['result'] = tb.cell(row=row,column=7).value   
        print('lis_item_nam--------')
        print(all_lis_item_name[i-1],all_lis_item_code[i-1] ,result_dct['result'] )
        # lis_item_name =  tb.cell(row=row,column=2).value
        result_dct['ext_combine_code'] = result_dct['lis_combine_code']
        
        result_info += set_result_info(result_dct)
    result_info = result_info + '<report_pic><pic_content>'  

    return result_info ,  dct



# 对每个检测项目的xml的result_info设置值
# 对每个检测项目的xml的result_info设置值
def set_result_info(dct):
    print('*'*20)
    print(dct)
    result_info = f'''    
    <result_info>
        <result_seq/>
        <ext_combine_code>{dct.get('ext_combine_code')}</ext_combine_code>
        <ext_combine_name>{dct.get('ext_combine_name')}</ext_combine_name>
        <ext_item_code>{dct.get('ext_item_code')}</ext_item_code>
        <ext_item_name>{dct.get('ext_item_name')}</ext_item_name>
        <result>        <![CDATA[{dct.get('result')}]]]]>>        <![CDATA[</result>
        <result_unit>        <![CDATA[{dct.get('result_unit')}]]]]>>   <![CDATA[</result_unit>
        <result_flag/>
        <result_reference>    <![CDATA[{dct.get('result_reference')}]]]]>>       <![CDATA[</result_reference>
        <result_date/>
        <result_intstrmt_name/>
        <result_test_method/>
        <result_suggestion/>
        <result_remark/>
        <lis_combine_code>{dct.get('lis_combine_code')}</lis_combine_code>
        <lis_combine_name>{dct.get('lis_combine_name')}</lis_combine_name>
        <lis_item_code>{dct.get('lis_item_code')}</lis_item_code>
        <lis_item_name>{dct.get('lis_item_name')}</lis_item_name>    
    </result_info>
   '''
    return result_info

# 要发送的xml的第一部分信息：
# import string 
def set_data1(dct):
    data1 = f'''
    <soap:Envelope
        xmlns:soap="http://www.w3.org/2003/05/soap-envelope"
        xmlns:tem="http://tempuri.org/">
        <soap:Header/>
        <soap:Body>
            <tem:UploadLisRepData>
                <!--Optional:-->
                <tem:reportResult>
                    <![CDATA[<Report_Result><Report_Info>
                    <ext_lab_code></ext_lab_code>
                    <lis_Barcode>{dct.get('lis_Barcode')}</lis_Barcode>
                    <ext_Barcode>{dct.get('ext_Barcode')}</ext_Barcode>
                    <ext_checkItem/>
                    <pat_name/>
                    <pat_age></pat_age>
                    <pat_height/>
                    <pat_wight/>
                    <pat_pre_week/>
                    <pat_id/>
                    <pat_bedNo/>
                    <pat_tel/>
                    <pat_sex/>
                    <pat_birthday/>
                    <pat_ori_name/>
                    <sam_name/>
                    <sam_state/>
                    <doctor_name/>
                    <dept_name/>
                    <clinical_diag/>
                    <SampleNumber/>
                    <blood_time>{dct.get('blood_time')}</blood_time>
                    <ext_check_ID/>
                    <ext_receive_time>{dct.get('ext_receive_time')}</ext_receive_time>
                    <ext_upload_time/>
                    <ext_report_suggestion/>
                    <ext_report_remark/>
                    <ext_intstrmt_name/>
                    <ext_lab_name/>
                    <ext_report_type/>
                    <ext_check_time/>
                    <ext_first_audit_time/>
                    <ext_first_audit_time/>
                    <ext_second_audit_time/>
                    <ext_checker/>
                    <ext_first_audit/>
                    <ext_second_audit/>
                    <ext_report_code>{dct.get('lis_Barcode')}</ext_report_code>
    '''
    return data1 


# pdf部分
def get_pdf(inpdf):
    data2 = ''
    with open(inpdf,'rb') as fh:
            data2 = base64.b64encode(fh.read())
    return data2

# 最后是将xml和pdf的信息发送到妇幼
def send_data(dct,data1,result_info,data2):
    # xml的结尾部分
    data3 = '''</pic_content><pic_name>{dct.get('lis_Barcode')}_report_base64.pdf</pic_name><pic_seq></pic_seq></report_pic></Report_Info></Report_Result>]]>
            </tem:reportResult>
        </tem:UploadLisRepData>
    </soap:Body>
</soap:Envelope>
    '''
    encode_data = data1.encode('utf-8') +result_info.encode('utf-8') + data2 +  data3.encode('utf-8')
    
    headers = {"Host": "10.10.11.196",
            "Content-Type": "text/xml; charset=utf-8",
            "Content-Length": str(len(encode_data)) ,#}
            "SOAPAction": "http://tempuri.org/UploadLisRepData"}
    patient_info = requests.post(f'http://58.62.17.237:4431/ExtReportService', data=encode_data ,headers=headers)

    print('patient_info.text----')    
    print(patient_info.text)

    if not os.path.exists('workdir'):
        os.mkdir('workdir')
    with open('./workdir/'+ dct.get('lis_Barcode') + '.xml','wb') as fh:
        fh.write(encode_data)


